#-*-coding:utf-8 -*-
# @Time    :2019\1\1 0001 {TIME}
# @Author  :LiHaiLang
# @Email   :1614461577@qq.com
# @File    :run_test.py
# @Software:PyCharm
import openpyxl
import json
from api_test2.common.Do_request import DoRequest
from api_test2.common import Do_contants
from api_test2.common.Do_config import Config
class Case:#对excel里面的case进行描述
    def __init__(self):
        self.case_id=None
        self.url=None
        self.data=None
        self.title=None
        self.http_method=None
        self.expected=None
        self.actual=None
        self.result=None

class DoExcel:
    def __init__(self,file_name):
        try:
            self.file_name=file_name
            self.workbook=openpyxl.load_workbook(filename=file_name)#打开excel
        except FileNotFoundError as e:
            print("未找到{0}文件，请检查".format(file_name))
            raise e
    def get_cases(self,sheet_name):
        sheet=self.workbook[sheet_name]#根据sheet_name获取此sheet下的所有数据
        max_row=sheet.max_row#获取最大行
        case_list=[]#创建一个实例来存放测试用例
        for i in range (2,max_row+1):
            case=Case()#实例化一个对象来存放数据
            case.case_id=sheet.cell(row=i,column=1).value#获取行列值
            case.url = sheet.cell(row=i, column=2).value
            case.data = sheet.cell(row=i, column=3).value
            case.title = sheet.cell(row=i, column=4).value
            case.http_method = sheet.cell(row=i, column=5).value
            case.expected = sheet.cell(row=i, column=6).value
            case_list.append(case)
        return case_list
    def get_sheet_names(self):
        return self.workbook.sheetnames
    def write_result(self,sheet_name,case_id,actual,result):#根据sheet_name获取到当前工作表，再根据case_id定位到行，
        #取到当前行里面的actual,result,赋值保存
        sheet=self.workbook[sheet_name]
        max_row=sheet.max_row#获取最大行
        for r in range(2,max_row+1):
            case_id_r=sheet.cell(r,1).value#取到第r行，第一列的值
            if case_id_r == case_id:#判断从excel里面取到的值是否与传入的case_id一致
                sheet.cell(r,7).value=actual
                sheet.cell(r,8).value=result
                self.workbook.save(filename=self.file_name)
                break
if __name__ == '__main__':
    run_excel=DoExcel("../datas/testcase.xlsx")#创建实例定位到excel测试用例
    sheetnames=run_excel.get_sheet_names()#获取所有的sheetnames
    print('sheet名称为：',sheetnames)
    case_lists=['register','login','recharge']#对不想跑的用例进行过滤

    for sheet in sheetnames:#从sheetnames里面每循环一次就取一个sheetname
        if sheet in case_lists:#如果当前sheet不在case_lists里面，就不执行
            cases = run_excel.get_cases(sheet)
            print(sheet+'测试用例个数：',len(cases))
            for case in cases:#遍历测试用例，每循环一次就取一条测试用例
                print('case信息：',case.__dict__)#打印case信息
                data=eval(case.data)#将测试用例里面的字典转换成字符串
                url=Config().get('api','url')+case.url
                res=DoRequest(method=case.http_method,url=url,data=data)#获取接口请求
                print('status_code:',res.get_status_code())#打印响应码
                res_dict=res.get_json()#获取请求响应，字典
                res_text=json.dumps(res_dict,ensure_ascii=False, indent=4)
                print(res_text)
                #判断excel里面的excepted是否与返回的响应一致
                if case.expected == res.get_text():
                    print('pass')
                    run_excel.write_result(sheet_name=sheet,case_id=case.case_id,actual=res_text,result='pass')
                else:
                    print('fail')
                    run_excel.write_result(sheet_name=sheet, case_id=case.case_id, actual=res_text, result='fail')












